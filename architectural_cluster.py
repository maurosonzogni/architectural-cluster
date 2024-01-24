import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from scipy.cluster.hierarchy import dendrogram, linkage, fcluster
from scipy.spatial.distance import squareform

from sklearn.metrics import silhouette_score

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, NamedStyle
from ai_modules.gtp_module import infer_topic_with_GPT
from cluster_module import build_cluster_contents_map, build_cluster_information, build_model_topics_chain_in_clusters, infer_cluster_topics

# local modules
from utils_module import generate_link, load_config, create_parent_folders

# Configuration file path
config_file_path = 'configurations/architectural_cluster_config.json'

# Load configuration
config = load_config(config_file_path)

# Load data from the CSV file
data = pd.read_csv(config['file_path'])

# Extract the similarity similarities from the dataframe
similarities = data.iloc[1:, :].values  # Exclude the "model_name" column
similarities = data.iloc[:, 1:].values  # Exclude the "model_name" column

# Convert the distance matrix into a condensed distance matrix
# This is because our matrix is already a distance matrix, but not in the required form
condensed_distance_matrix = squareform(similarities)

# The hierarchy.linkage function from the scipy.cluster module is used to compute a linkage matrix from a distance matrix or a condensed distance vector.
# This linkage matrix is later used to construct the dendrogram, representing the hierarchical structure of the clustering.
# Z = hierarchy.linkage(y, method='single', metric='euclidean')
#   y: Distance matrix or condensed distance vector.
#   method: Linkage method to use. It can be one of the following: 'single', 'complete', 'average', 'weighted', 'centroid', 'median', 'ward'.
#           These methods determine how the distance between clusters is calculated from distances of individual points within clusters.
#   metric: Specifies the distance metric to use. It can be a string representing one of the distance metrics supported by scipy.spatial.distance.pdist.
#           The default metric is 'euclidean'.
# The function returns the linkage matrix Z, which is an (n-1) x 4 matrix, where n is the number of observations.
# Each row of this matrix represents a fusion between two clusters and contains the following information:
#   Index of the first cluster.
#   Index of the second cluster.
#   Distance between the first and second cluster.
#   Number of observations in the resulting merged cluster.

# The dendrogram is then constructed using this linkage matrix and
# visualizes the sequence of cluster fusions during the hierarchical clustering process.

# Calculate linkage matrix using average linkage method
Z = linkage(condensed_distance_matrix, 'average')

# Create a new figure
plt.figure()

# Set the window title for the dendrogram
plt.gcf().canvas.get_tk_widget().master.title('Dendrogram')

# Set the title for the plot
plt.title('Models Cluster')

# Adjust subplot parameters for better layout
plt.subplots_adjust(left=0.08, right=0.950, bottom=0.435, top=0.935)
# Traccia una linea orizzontale all'altezza del taglio
plt.axhline(y=config['cluster_cut_height'], color='r', linestyle='--')

# Create a dendrogram with hierarchical clustering labels
dn = dendrogram(Z, labels=data['model_name'].values)

cluster_contents_map = build_cluster_contents_map(
    Z, data['model_name'].values, config['cluster_cut_height'], config['min_cluster_size'], config['max_cluster_size'])

cluster_info_list = build_cluster_information(
    cluster_contents_map, config['method_to_infer_topics'], config['numeber_of_topic_to_infer'], config['common_words_to_exclude'])

# Utilizzo:
# Assumi che `cluster_info_list` sia già definito come discusso in precedenza
model_to_topics_map = build_model_topics_chain_in_clusters(cluster_info_list)

# Assign cluster labels using hierarchical clustering
labels = fcluster(Z, t=config['cluster_cut_height'], criterion='distance')

# Initialize a dictionary to store models assigned to each cluster

clusters = cluster_contents_map

# Calculate silhouette score for all samples together
# Note: To calculate silhouette for individual clusters, each cluster must contain at least 2 labels
silhouette_avg = silhouette_score(similarities, labels, metric='precomputed')

# Create a DataFrame for metrics
metrics_data = [[config['file_path'], config['cluster_cut_height'],
                 silhouette_avg, config['numeber_of_topic_to_infer']]]
metrics_df = pd.DataFrame(metrics_data, columns=[
                          'file_path', 'cluster_cut_height', 'silhouette_avg', 'numeber_of_topic_to_infer'])

# Initialize a list to store cluster information
cluster_data = []

method_to_infer_topics_not_found = True

# NOTE Python does not have a native switch construction. If one wishes to add a different mode, add a custom if condition.
if config['method_to_infer_topics'] == 'GPT':
    method_to_infer_topics_not_found = False
    # Populate the list with cluster information, including representative labels
    for cluster_num, cluster_labels in clusters.items():
        cluster_topics = infer_topic_with_GPT(
            cluster_labels['members'], "", "", 'user')
        cluster_data.append([cluster_topics, cluster_num,
                            cluster_labels['merged'], cluster_labels['members']])

# Both, native and default case
if method_to_infer_topics_not_found or config['method_to_infer_topics'] == 'NATIVE':
    # Populate the list with cluster information, including representative labels
    for cluster_num, cluster_labels in clusters.items():
        cluster_topics = infer_cluster_topics(
            cluster_labels['members'], config['numeber_of_topic_to_infer'], config['common_words_to_exclude'])
        cluster_data.append([cluster_topics, cluster_num,
                            cluster_labels['merged'], cluster_labels['members']])


# Create a DataFrame for cluster information
cluster_df = pd.DataFrame(cluster_data, columns=[
                          'cluster_topic', 'cluster_number', 'merged_clusters', 'contained_models_name'])

# Trova il massimo numero di colonne tra tutte le liste
max_num_colonne = 0

for item, content in clusters.items():
    if (max_num_colonne < len(content['members'])):
        max_num_colonne = len(content['members'])

for i in range(max_num_colonne):
    colonna_label = f'model_{i+1}'
    cluster_df[colonna_label] = [generate_link(list(labels['members'])[i]) if i < len(
        labels['members']) else '' for labels in clusters.values()]


# Create the folder structure if it doesn't exist
create_parent_folders(config['clusters_output_xlsx'])

# Write the silhouette score DataFrame to Excel
metrics_df.to_excel(config['clusters_output_xlsx'],
                    sheet_name=config['metrics_sheet_name'], index=False)

# Convertendo la lista di dizionari in un DataFrame
model_cluster_chain_df = pd.DataFrame(model_to_topics_map)

# Convertendo la colonna 'topics' in stringa per una migliore visualizzazione in Excel
model_cluster_chain_df['topics'] = model_cluster_chain_df['topics'].apply(
    lambda x: ', '.join(x))


# Crea uno stile per l'attributo hyperlink
hyperlink_style = NamedStyle(name='hyperlink', font=Font(
    underline='single', color='0563C1'))
# Append the cluster information DataFrame to the existing Excel file
with pd.ExcelWriter(config['clusters_output_xlsx'], engine='openpyxl', mode='a') as writer:

    model_cluster_chain_df.to_excel(
        writer, sheet_name=config['model_cluster_chain_sheet_name'], index=False)

    cluster_df.to_excel(
        writer, sheet_name=config['cluster_sheet_name'], index=False)

    ws = writer.sheets[config['cluster_sheet_name']]

    # Imposta le colonne da trasformare in link (inizia dalla quinta colonna)
    colonne_da_trasformare = cluster_df.columns[4:]
    # Itera attraverso le colonne e le celle per aggiungere l'attributo hyperlink

    # Itera attraverso le colonne e le celle per aggiungere l'attributo hyperlink
    for colonna in colonne_da_trasformare:
        indice_colonna = cluster_df.columns.get_loc(colonna) + 1
        for col in ws.iter_cols(min_col=indice_colonna, max_col=indice_colonna, min_row=2):
            for indice, cella in enumerate(col, start=2):
                cella.style = hyperlink_style
                valore_cella = cella.value
                valore_ipo = None if valore_cella is None else f'=HYPERLINK("{valore_cella}", "{valore_cella}")'
                ws[f'{get_column_letter(indice_colonna)}{indice}'] = valore_ipo

    # Imposta la larghezza automatica delle colonne
    for col_num, _ in enumerate(cluster_df.columns, 1):
        col_letter = get_column_letter(col_num)
        max_length = 0
        for row in ws[f'{col_letter}']:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(row.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width


print(
    f"\nThe results have been saved in the Excel file: {config['clusters_output_xlsx']}")

# Display the dendrogram and silhouette analysis
plt.show()
